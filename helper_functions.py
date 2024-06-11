from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from typing import Any, Union

import requests
from io import BytesIO
import fitz
import pandas as pd

from openai import OpenAI
from openai import AuthenticationError
import os
from dotenv import load_dotenv
from typing import List
import numpy as np

load_dotenv()


def get_sheet_from_excel(filename: Path, sheet_name: str) -> Worksheet:
    """
    Get a worksheet from an Excel file

    Args:
    filename (Path): The path to the Excel file
    sheet_name (str): The name of the worksheet

    Returns:
    Worksheet: The worksheet object
    """
    wb = load_workbook(filename)
    return wb[sheet_name]


def serialize_value(cell: Cell) -> str:
    value = cell.value
    return str(value)


def remove_none_key_value_pairs(d: dict[Any, Any]) -> dict[Any, Any]:
    """
    Remove key-value pairs where both the key and value are None

    Returns:
    dict: A new dictionary with None key-value pairs removed.
    """

    return {
        key: value for key, value in d.items() if not (key is None and value is None)
    }


def process_simple_table(ws: Worksheet) -> list[dict[str, Union[str, float, int]]]:
    """
    process_simple_table handles a simple spreadsheet which has one table starting from the top left corner
    Its first row is its header and the following rows are data records.
    Example:
    | Month    | Savings |
    | -------- | ------- |
    | January  | $250    |
    | February | $80     |
    | March    | $420    |
    """
    headers = [serialize_value(cell) for cell in ws[1]]

    records = []
    for row in ws.iter_rows(min_row=2):
        values = [serialize_value(cell) for cell in row]
        record = dict(zip(headers, values))
        records.append(remove_none_key_value_pairs(record))
    return records


def calculate_num_leading_space_per_level(row_headers: list[str]) -> int:
    for current_header, next_header in zip(row_headers, row_headers[1:]):
        current_spaces = len(current_header) - len(current_header.lstrip())
        next_spaces = len(next_header) - len(next_header.lstrip())
        if next_spaces != current_spaces:
            return next_spaces - current_spaces
    return 0


def process_hierarchical_table(ws: Worksheet) -> dict[str, Any]:
    """
    process_hierarchical_table handles a spreadsheet which has one table starting from the top left corner
    Its top left cell is empty. Its first row and first column are its headers.
    Its first column has hierarchical structural represented by the number of leading spaces.
    Some rows represents a category where the data cells are empty. Other rows represents actual data where data can be found in the data cells.
    Example:
    |                                              |30-Sep-23           |31-Oct-23           |30-Nov-23           |
    |----------------------------------------------|--------------------|--------------------|--------------------|
    |Assets                                        |                    |                    |                    |
    |   Current Assets                             |                    |                    |                    |
    |      Cash and Cash Equivalent                |                    |                    |                    |
    |         1060 TD Chequing Bank Account - #4092|587,881.66          |750,736.21          |453,234.78          |
    |         1061 TD AUD FX Currency-XXX-0283     |1,588.43            |17,457.51           |1,444.33            |
    |      Total Cash and Cash Equivalent          |$         589,470.09|$         768,193.72|$         454,679.11|
    |      1320 Prepaid Expenses                   |423,826.69          |233,127.50          |270,189.85          |
    |         1302 Prepaid License at Vid Australia|46,985.98           |68,985.98           |68,985.98           |
    |   Total Current Assets at Inc. and Australia |$      1,060,282.76 |$      1,070,307.20 |$         793,854.94|
    |Total Assets                                  |$      1,060,282.76 |$      1,070,307.20 |$         793,854.94|
    """

    def add_data(
        processed_table: dict[str, Any],
        nodes: list[str],
        col_headers: list[str],
        data_cells: tuple[Cell, ...],
    ) -> dict[str, Any]:
        current_level = processed_table
        for node in nodes[:-1]:
            if node not in current_level:
                print(
                    f"warning: can't find node {node} in processed table {current_level}. Creating a new node."
                )
                current_level[node] = {}
            current_level = current_level[node]

        current_level[nodes[-1]] = dict(
            zip(col_headers, [serialize_value(d) for d in data_cells])
        )
        return processed_table

    col_headers = [serialize_value(e) for e in ws[1][1:]]

    row_headers: list[str] = []
    for column in ws.iter_cols(min_col=1, max_col=1, values_only=False):
        row_headers = [serialize_value(cell) for cell in column[1:]]

    num_leading_space_per_level = calculate_num_leading_space_per_level(row_headers)

    if num_leading_space_per_level == 0:
        num_leading_space_per_level = 1

    processed_table: dict[str, Any] = {}
    nodes: list[str] = []

    # Process each row into the hierarchical structure
    for row in ws.iter_rows(min_row=2, values_only=False):
        level = (len(serialize_value(row[0])) - len(serialize_value(row[0]).lstrip())) // num_leading_space_per_level
        label = serialize_value(row[0]).strip()
        data_cells = row[1:]

        nodes = nodes[:level]
        nodes.append(label)

        if any([c for c in data_cells if c.value is not None]):
            processed_table = add_data(processed_table, nodes, col_headers, data_cells)

    return remove_none_key_value_pairs(processed_table)


def fetch_document(source):
    if source.startswith('http://') or source.startswith('https://'):
        response = requests.get(source)
        response.raise_for_status()
        return BytesIO(response.content)
    elif os.path.exists(source):
        return source
    else:
        raise FileNotFoundError(f"No such file or URL: {source}")
    
def extract_text_from_csv(csv_file):
    df = pd.read_csv(csv_file)
    return ' '.join(df.fillna('').astype(str).values.flatten())

def extract_text_from_pdf(pdf_file):
    doc = fitz.open(stream=pdf_file, filetype="pdf")
    text = ""
    for page in doc:
        text += page.get_text()
    return text

def give_serialized_string(excel_file):
    tables = serialize_excel_tables(excel_file=excel_file)
    serialized_tables = []
    for table in tables:
        table_text = '\n'.join(['\t'.join(row) for row in table])
        serialized_tables.append(table_text)
    
    result_text = '\n\n'.join(serialized_tables)
    
    return result_text

def serialize_excel_tables(file):
    # Load workbook and get active sheet
    wb = load_workbook(file, data_only=True)
    sheet = wb.active

    # Helper functions for border detection
    def has_border(cell: Cell) -> bool:
        return has_row_border(cell) or has_col_border(cell)

    def has_row_border(cell: Cell) -> bool:
        border = cell.border
        return bool(border.top.style or border.bottom.style)

    def has_col_border(cell: Cell) -> bool:
        border = cell.border
        return bool(border.left.style or border.right.style)

    # Convert the sheet into a 2D list of cells and flags for processing
    rows = list(sheet.iter_rows())
    flags = np.zeros((len(rows), len(rows[0])), dtype=int)
    
    tables = []

    def get_table(start_row_idx: int, start_col_idx: int) -> List[List[Any]]:
        header_row = rows[start_row_idx]
        end_col_idx = len(header_row) - 1
        
        for col_idx in range(start_col_idx, len(header_row)):
            cell = header_row[col_idx]
            if not has_border(cell):
                end_col_idx = col_idx - 1
                break

        table = []
        for row_idx in range(start_row_idx, len(rows)):
            cell = rows[row_idx][start_col_idx]
            if not has_border(cell):
                break
            row_data = []
            for col_idx in range(start_col_idx, end_col_idx + 1):
                row_data.append(rows[row_idx][col_idx].value)
                flags[row_idx][col_idx] = 1
            table.append(row_data)
        return table

    # Iterate over all cells to find tables based on borders
    for row_idx, row in enumerate(rows):
        for col_idx, cell in enumerate(row):
            if flags[row_idx][col_idx] == 0 and has_border(cell):
                table = get_table(row_idx, col_idx)
                if table:
                    tables.append(table)

    # if tables have a table in which every item of a row is null then remove that row
    for table in tables:
        for row in table:
            if all([cell is None for cell in row]):
                table.remove(row)
    return tables


def clean_text(text):
    return ' '.join(text.split())

def get_openai_client(api_key):
    try:
        client = OpenAI(api_key=api_key)
        # Perform a minimal request to validate the API key
        client.models.list()
        return client
    
    except AuthenticationError as e:
        raise ValueError("Invalid OpenAI API key.") from e
    
    except Exception as e:
        raise

def answer_question(text, question):
    openai_api_key = os.getenv('OPENAI_API_KEY')
    openai_client = get_openai_client(openai_api_key)

    completion = openai_client.chat.completions.create(
        model="gpt-4",
        messages=[
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": text},
            {"role": "user", "content": question}
        ]
    )
    return completion.choices[0].message.content